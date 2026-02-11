import re
import pdfplumber
import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
from collections import defaultdict
from datetime import datetime
import io
import json
import base64
import os
import hashlib
from pathlib import Path

PRECIO_BN = 0.0098
PRECIO_COLOR = 0.119
IVA = 0.21

# Directorio para almacenamiento persistente
DATA_DIR = Path("factubam_data")
DATA_DIR.mkdir(exist_ok=True)
HISTORIAL_FILE = DATA_DIR / "historial.json"
DOCUMENTOS_DIR = DATA_DIR / "documentos"
DOCUMENTOS_DIR.mkdir(exist_ok=True)

st.set_page_config(
    page_title="FactuBAM",
    page_icon="Imagenes/ada-logo.png",
    layout="wide"
)
try:
    st.image(
        "Imagenes/cabecera_andalucia.jpg",
        use_container_width=True
    )
except:
    pass # Evitar error si no existe la imagen

st.markdown("<br><br>", unsafe_allow_html=True)

# ======================================================
# ESTILOS CORPORATIVOS (COLORES Y TIPOGRAFÍAS)
# ======================================================
st.markdown("""
<style>
/* Paleta Junta de Andalucía */
:root {
    --verde-principal: #007932;
    --verde-secundario: #368F3F;
    --negro-corporativo: #2E2D25;
    --gris-claro: #D9D9D9;
}

/* Tipografía corporativa (si está disponible en Tipografias/) */
@font-face {
    font-family: 'JuntaSans';
    src: url('Tipografias/JuntaSans-Regular.ttf') format('truetype');
}

/* Aplicación global */
html, body, [class*="css"] {
    font-family: 'JuntaSans', Arial, sans-serif;
}

/* Títulos */
h1, h2, h3 {
    color: var(--verde-principal);
}

/* Botones */
.stButton > button {
    background-color: var(--verde-principal);
    color: white;
    border-radius: 4px;
    border: none;
}

.stButton > button:hover {
    background-color: var(--verde-secundario);
}

/* Métricas */
[data-testid="stMetricValue"] {
    color: var(--verde-principal);
}

/* Expander */
details summary {
    background-color: var(--gris-claro);
    color: var(--negro-corporativo);
    padding: 6px;
    border-radius: 4px;
}
</style>
""", unsafe_allow_html=True)

# Funciones de almacenamiento persistente en disco
def cargar_historial():
    """Carga el historial desde archivo JSON local"""
    try:
        if not HISTORIAL_FILE.exists():
            return []
        
        with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
            historial_data = json.load(f)
        
        historial = []
        for reg_data in historial_data:
            try:
                # Cargar DataFrame
                df_file = DOCUMENTOS_DIR / f"{reg_data['id']}_data.json"
                if df_file.exists():
                    with open(df_file, 'r', encoding='utf-8') as f:
                        df_data = json.load(f)
                        reg_data['df'] = pd.DataFrame(df_data)
                
                # Cargar archivos PDF y Excel
                pdf_file = DOCUMENTOS_DIR / f"{reg_data['id']}_factura.pdf"
                excel_file = DOCUMENTOS_DIR / f"{reg_data['id']}_inventario.xlsx"
                
                if pdf_file.exists():
                    with open(pdf_file, 'rb') as f:
                        reg_data['pdf_bytes'] = f.read()
                
                if excel_file.exists():
                    with open(excel_file, 'rb') as f:
                        reg_data['excel_bytes'] = f.read()
                
                historial.append(reg_data)
            except Exception as e:
                st.warning(f"Error al cargar registro {reg_data.get('id', 'desconocido')}: {str(e)}")
                continue
        
        return historial
    except Exception as e:
        st.error(f"Error al cargar historial: {str(e)}")
        return []

def guardar_historial(historial):
    """Guarda el historial en archivo JSON local"""
    try:
        # Preparar datos para JSON (sin DataFrame ni bytes)
        historial_simple = []
        
        for registro in historial:
            reg_simple = {
                'id': registro['id'],
                'nombre': registro['nombre'],
                'fecha_hora': registro['fecha_hora'],
                'pdf_name': registro['pdf_name'],
                'excel_name': registro['excel_name'],
                'dispositivos': registro['dispositivos'],
                'coste_total_sin_iva': registro['coste_total_sin_iva'],
                'coste_total_con_iva': registro['coste_total_con_iva']
            }
            historial_simple.append(reg_simple)
            
            # Guardar DataFrame como JSON separado
            df_file = DOCUMENTOS_DIR / f"{registro['id']}_data.json"
            with open(df_file, 'w', encoding='utf-8') as f:
                json.dump(registro['df'].to_dict('records'), f, ensure_ascii=False, indent=2)
            
            # Guardar archivos PDF y Excel
            if 'pdf_bytes' in registro:
                pdf_file = DOCUMENTOS_DIR / f"{registro['id']}_factura.pdf"
                with open(pdf_file, 'wb') as f:
                    f.write(registro['pdf_bytes'])
            
            if 'excel_bytes' in registro:
                excel_file = DOCUMENTOS_DIR / f"{registro['id']}_inventario.xlsx"
                with open(excel_file, 'wb') as f:
                    f.write(registro['excel_bytes'])
        
        # Guardar índice principal
        with open(HISTORIAL_FILE, 'w', encoding='utf-8') as f:
            json.dump(historial_simple, f, ensure_ascii=False, indent=2)
        
        return True
    except Exception as e:
        st.error(f"Error al guardar historial: {str(e)}")
        return False

def eliminar_registro_disco(registro_id):
    """Elimina los archivos de un registro del disco"""
    try:
        # Eliminar archivos del registro
        archivos = [
            DOCUMENTOS_DIR / f"{registro_id}_data.json",
            DOCUMENTOS_DIR / f"{registro_id}_factura.pdf",
            DOCUMENTOS_DIR / f"{registro_id}_inventario.xlsx"
        ]
        
        for archivo in archivos:
            if archivo.exists():
                archivo.unlink()
        
        return True
    except Exception as e:
        st.error(f"Error al eliminar archivos: {str(e)}")
        return False

def limpiar_historial_disco():
    """Limpia todo el historial del disco"""
    try:
        # Eliminar todos los archivos en el directorio de documentos
        for archivo in DOCUMENTOS_DIR.glob("*"):
            archivo.unlink()
        
        # Eliminar archivo de historial
        if HISTORIAL_FILE.exists():
            HISTORIAL_FILE.unlink()
        
        return True
    except Exception as e:
        st.error(f"Error al limpiar historial: {str(e)}")
        return False

# Inicializar session_state
if 'historial_documentos' not in st.session_state:
    # Cargar desde disco
    st.session_state.historial_documentos = cargar_historial()

if 'registro_seleccionado' not in st.session_state:
    st.session_state.registro_seleccionado = None
if 'mostrar_nuevo' not in st.session_state:
    st.session_state.mostrar_nuevo = True
if 'modo_vista' not in st.session_state:
    st.session_state.modo_vista = 'nuevo'
if 'documentos_seleccionados' not in st.session_state:
    st.session_state.documentos_seleccionados = []

def extraer_datos_pdf(pdf_bytes):
    datos = defaultdict(lambda: {"bn": 0, "color": 0})
    sn_actual = None
    with pdfplumber.open(pdf_bytes) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if not tables:
                continue
            for table in tables:
                for fila in table:
                    if not fila or len(fila) < 3:
                        continue
                    desc = str(fila[1]).upper()
                    cantidad = fila[2]
                    match_sn = re.search(r'([A-Z0-9]{8,})\s+N/S', desc)
                    if match_sn:
                        sn_actual = match_sn.group(1)
                        continue
                    if sn_actual is None:
                        continue
                    if "TOTAL MONOCROMO" in desc:
                        try:
                            datos[sn_actual]["bn"] = int(float(str(cantidad).replace('.', '').replace(',', '.')))
                        except:
                            datos[sn_actual]["bn"] = 0
                    if "TOTAL COLOR" in desc:
                        try:
                            datos[sn_actual]["color"] = int(float(str(cantidad).replace('.', '').replace(',', '.')))
                        except:
                            datos[sn_actual]["color"] = 0
    return datos

def cruzar_excel(xlsx_file, datos_pdf):
    wb = openpyxl.load_workbook(xlsx_file)
    resultados = []
    
    # IMPORTANTE: Creamos un conjunto para rastrear qué números de serie del PDF 
    # encontramos en el Excel. Así sabremos cuáles faltan.
    sns_encontrados_en_excel = set()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        header = [c.value for c in sheet[1]]
        if "S/N" not in header:
            continue
        idx_sn = header.index("S/N") + 1
        idx_org = header.index("Organismo") + 1
        idx_ubi = header.index("Ubicación exacta") + 1
        for row in range(2, sheet.max_row + 1):
            sn = sheet.cell(row, idx_sn).value
            
            # Limpieza básica del SN del Excel para asegurar coincidencia
            if sn:
                sn = str(sn).strip()

            if sn in datos_pdf:
                # Marcamos este SN como encontrado
                sns_encontrados_en_excel.add(sn)

                bn = datos_pdf[sn]["bn"]
                color = datos_pdf[sn]["color"]
                
                coste_bn_sin_iva = bn * PRECIO_BN
                coste_color_sin_iva = color * PRECIO_COLOR
                coste_sin_iva = coste_bn_sin_iva + coste_color_sin_iva
                
                iva_bn = coste_bn_sin_iva * IVA
                iva_color = coste_color_sin_iva * IVA
                iva_total = iva_bn + iva_color
                
                coste_bn_con_iva = coste_bn_sin_iva + iva_bn
                coste_color_con_iva = coste_color_sin_iva + iva_color
                coste_con_iva = coste_sin_iva + iva_total
                
                resultados.append({
                    "sn": sn,
                    "organismo": sheet.cell(row, idx_org).value,
                    "ubicacion": sheet.cell(row, idx_ubi).value,
                    "bn": bn,
                    "color": color,
                    "coste_bn_sin_iva": coste_bn_sin_iva,
                    "coste_color_sin_iva": coste_color_sin_iva,
                    "coste_sin_iva": coste_sin_iva,
                    "iva_bn": iva_bn,
                    "iva_color": iva_color,
                    "iva_total": iva_total,
                    "coste_bn_con_iva": coste_bn_con_iva,
                    "coste_color_con_iva": coste_color_con_iva,
                    "coste_con_iva": coste_con_iva,
                    "estado": "Revisado"
                })
    
    # === CORRECCIÓN DEL DESCUADRE ===
    # Buscar qué S/N existen en el PDF pero NO se encontraron en el Excel
    # Estos son los que provocan la diferencia de precio.
    for sn_pdf, valores in datos_pdf.items():
        if sn_pdf not in sns_encontrados_en_excel:
            # Calcular costes de la máquina "huérfana"
            bn = valores["bn"]
            color = valores["color"]
            
            coste_bn_sin_iva = bn * PRECIO_BN
            coste_color_sin_iva = color * PRECIO_COLOR
            coste_sin_iva = coste_bn_sin_iva + coste_color_sin_iva
            
            iva_bn = coste_bn_sin_iva * IVA
            iva_color = coste_color_sin_iva * IVA
            iva_total = iva_bn + iva_color
            
            coste_bn_con_iva = coste_bn_sin_iva + iva_bn
            coste_color_con_iva = coste_color_sin_iva + iva_color
            coste_con_iva = coste_sin_iva + iva_total

            # Añadir al resultado con un aviso visible
            resultados.append({
                "sn": sn_pdf,
                "organismo": "⚠️ NO EN EXCEL (Solo Factura)",
                "ubicacion": "Desconocida",
                "bn": bn,
                "color": color,
                "coste_bn_sin_iva": coste_bn_sin_iva,
                "coste_color_sin_iva": coste_color_sin_iva,
                "coste_sin_iva": coste_sin_iva,
                "iva_bn": iva_bn,
                "iva_color": iva_color,
                "iva_total": iva_total,
                "coste_bn_con_iva": coste_bn_con_iva,
                "coste_color_con_iva": coste_color_con_iva,
                "coste_con_iva": coste_con_iva,
                "estado": "⚠️ Faltante en Excel"
            })

    return resultados

def guardar_registro(nombre, pdf_file, excel_file, df):
    """Guarda un registro completo con los datos procesados"""
    pdf_bytes = pdf_file.read()
    pdf_file.seek(0)
    excel_bytes = excel_file.read()
    excel_file.seek(0)
    
    nuevo_id = int(datetime.now().timestamp() * 1000)
    
    registro = {
        'id': nuevo_id,
        'nombre': nombre,
        'fecha_hora': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'pdf_name': pdf_file.name,
        'excel_name': excel_file.name,
        'pdf_bytes': pdf_bytes,
        'excel_bytes': excel_bytes,
        'df': df.copy(),
        'dispositivos': len(df),
        'coste_total_sin_iva': df['coste_sin_iva'].sum(),
        'coste_total_con_iva': df['coste_con_iva'].sum()
    }
    st.session_state.historial_documentos.append(registro)
    guardar_historial(st.session_state.historial_documentos)

def eliminar_registro(registro_id):
    """Elimina un registro del historial y del disco"""
    # Eliminar del disco
    eliminar_registro_disco(registro_id)
    
    # Eliminar de la memoria
    st.session_state.historial_documentos = [
        r for r in st.session_state.historial_documentos if r['id'] != registro_id
    ]
    
    # Guardar historial actualizado
    guardar_historial(st.session_state.historial_documentos)

def limpiar_historial():
    """Limpia todo el historial del disco y memoria"""
    # Limpiar disco
    limpiar_historial_disco()
    
    # Limpiar memoria
    st.session_state.historial_documentos = []

def obtener_dataframe_acumulado(ids_seleccionados=None):
    """
    Combina los dataframes de múltiples registros
    Si ids_seleccionados es None, usa todos los registros
    """
    if ids_seleccionados is None:
        registros = st.session_state.historial_documentos
    else:
        registros = [r for r in st.session_state.historial_documentos if r['id'] in ids_seleccionados]
    
    if not registros:
        return None
    
    dfs = []
    for registro in registros:
        df_temp = registro['df'].copy()
        df_temp['documento'] = registro['nombre']
        df_temp['fecha'] = registro['fecha_hora']
        dfs.append(df_temp)
    
    df_acumulado = pd.concat(dfs, ignore_index=True)
    return df_acumulado

def mostrar_analisis(df, titulo="Análisis", mostrar_por_documento=False):
    """Muestra todas las gráficas y tablas del análisis"""
    
    st.subheader(titulo)
    
    # Métricas generales
    col1, col2, col3, col4, col5 = st.columns(5)
    
    if mostrar_por_documento:
        num_documentos = df['documento'].nunique()
        col1.metric("📄 Documentos", num_documentos)
    else:
        col1.metric("🖥️ Dispositivos", len(df))
    
    # Calcular dispositivos faltantes (los que añadimos nosotros)
    sin_ubicar = len(df[df['estado'] == "⚠️ Faltante en Excel"])
    revisados = len(df) - sin_ubicar

    col2.metric("✅ Revisados", revisados)
    col3.metric("❌ Sin ubicar (Solo PDF)", sin_ubicar) # Actualizado para mostrar el error
    col4.metric("🖨️ Total B/N", f"{df['bn'].sum():,}")
    col5.metric("🎨 Total Color", f"{df['color'].sum():,}")
    
    # Totales con y sin IVA
    col_iva1, col_iva2, col_iva3 = st.columns(3)
    col_iva1.metric("💰 Total sin IVA", f"{df['coste_sin_iva'].sum():,.2f} €")
    col_iva3.metric("💳 Total con IVA", f"{df['coste_con_iva'].sum():,.2f} €")
    
    st.markdown("---")
    
    # Tabs principales
    if mostrar_por_documento:
        tabs = st.tabs(["📄 Por Documento", "🏢 Por Departamento", "🔍 Detalle Equipos"])
        
        with tabs[0]:
            mostrar_analisis_por_documento(df)
        
        with tabs[1]:
            mostrar_analisis_por_departamento(df)
        
        with tabs[2]:
            mostrar_detalle_equipos(df)
    else:
        tabs = st.tabs(["🏢 Por Departamento", "🔍 Detalle Equipos"])
        
        with tabs[0]:
            mostrar_analisis_por_departamento(df)
        
        with tabs[1]:
            mostrar_detalle_equipos(df)

def mostrar_analisis_por_documento(df):
    """Análisis agrupado por documento"""
    df_docs = df.groupby('documento').agg({
        'bn': 'sum',
        'color': 'sum',
        'coste_sin_iva': 'sum',
        'coste_con_iva': 'sum',
        'iva_total': 'sum',
        'sn': 'count',
        'fecha': 'first'
    }).rename(columns={'sn': 'dispositivos'}).reset_index()
    
    df_docs['total_impresiones'] = df_docs['bn'] + df_docs['color']
    df_docs = df_docs.sort_values('fecha')
    
    st.markdown("### 📊 Resumen por Documento")
    
    # Gráfico de evolución de costes
    fig_evol = go.Figure()
    fig_evol.add_trace(go.Bar(
        name='Sin IVA',
        x=df_docs['documento'],
        y=df_docs['coste_sin_iva'],
        marker_color='lightblue'
    ))
    fig_evol.add_trace(go.Bar(
        name='Con IVA',
        x=df_docs['documento'],
        y=df_docs['coste_con_iva'],
        marker_color='darkblue'
    ))
    fig_evol.update_layout(
        title='Evolución de Costes por Documento',
        barmode='group',
        xaxis_title='Documento',
        yaxis_title='Coste (€)',
        height=500
    )
    st.plotly_chart(fig_evol, use_container_width=True, key="evol_costes_doc")
    
    # Gráfico de impresiones por documento
    fig_imp = go.Figure()
    fig_imp.add_trace(go.Scatter(
        name='B/N',
        x=df_docs['documento'],
        y=df_docs['bn'],
        mode='lines+markers',
        line=dict(color='gray', width=2)
    ))
    fig_imp.add_trace(go.Scatter(
        name='Color',
        x=df_docs['documento'],
        y=df_docs['color'],
        mode='lines+markers',
        line=dict(color='skyblue', width=2)
    ))
    fig_imp.update_layout(
        title='Evolución de Impresiones por Documento',
        xaxis_title='Documento',
        yaxis_title='Número de Impresiones',
        height=400
    )
    st.plotly_chart(fig_imp, use_container_width=True, key="evol_imp_doc")
    
    # Tabla resumen
    st.markdown("### 📋 Tabla Resumen por Documento")
    df_docs_display = df_docs.copy()
    df_docs_display['coste_sin_iva'] = df_docs_display['coste_sin_iva'].apply(lambda x: f"{x:.2f} €")
    df_docs_display['iva_total'] = df_docs_display['iva_total'].apply(lambda x: f"{x:.2f} €")
    df_docs_display['coste_con_iva'] = df_docs_display['coste_con_iva'].apply(lambda x: f"{x:.2f} €")
    df_docs_display['bn'] = df_docs_display['bn'].apply(lambda x: f"{x:,}")
    df_docs_display['color'] = df_docs_display['color'].apply(lambda x: f"{x:,}")
    df_docs_display['total_impresiones'] = df_docs_display['total_impresiones'].apply(lambda x: f"{x:,}")
    
    st.dataframe(
        df_docs_display[['documento', 'fecha', 'dispositivos', 'bn', 'color', 
                         'total_impresiones', 'coste_sin_iva', 'iva_total', 'coste_con_iva']].rename(columns={
            'documento': 'Documento',
            'fecha': 'Fecha',
            'dispositivos': 'Dispositivos',
            'bn': 'B/N',
            'color': 'Color',
            'total_impresiones': 'Total Impresiones',
            'coste_sin_iva': 'Coste sin IVA',
                        'coste_con_iva': 'Coste con IVA'
        }),
        use_container_width=True
    )

def mostrar_analisis_por_departamento(df):
    """Análisis agrupado por departamento"""
    df_dept = df.groupby('organismo').agg({
        'bn': 'sum',
        'color': 'sum',
        'coste_sin_iva': 'sum',
        'coste_con_iva': 'sum',
        'iva_total': 'sum',
        'sn': 'count'
    }).rename(columns={'sn': 'dispositivos'}).reset_index()
    
    df_dept['total_impresiones'] = df_dept['bn'] + df_dept['color']
    
    st.markdown("### 📊 Análisis por Departamento")
    
    tab1, tab2, tab3, tab4 = st.tabs(["💰 Coste", "🖨️ Impresiones", "🖥️ Dispositivos", "📋 Detalle"])
    
    with tab1:
        fig_coste_comp = go.Figure()
        fig_coste_comp.add_trace(go.Bar(
            name='Sin IVA',
            x=df_dept['organismo'],
            y=df_dept['coste_sin_iva'],
            marker_color='lightcoral'
        ))
        fig_coste_comp.add_trace(go.Bar(
            name='Con IVA',
            x=df_dept['organismo'],
            y=df_dept['coste_con_iva'],
            marker_color='darkred'
        ))
        fig_coste_comp.update_layout(
            title='Comparativa de Costes: Sin IVA vs Con IVA',
            barmode='group',
            xaxis_title='Departamento',
            yaxis_title='Coste (€)',
            height=500
        )
        st.plotly_chart(fig_coste_comp, use_container_width=True, key="dept_coste_comp")
        
        fig_pie_coste = px.pie(
            df_dept,
            values='coste_con_iva',
            names='organismo',
            title='Distribución de Costes por Departamento (con IVA)',
            hole=0.4
        )
        st.plotly_chart(fig_pie_coste, use_container_width=True, key="dept_pie_coste")
    
    with tab2:
        fig_impresiones = go.Figure()
        fig_impresiones.add_trace(go.Bar(
            name='B/N',
            x=df_dept['organismo'],
            y=df_dept['bn'],
            marker_color='lightgray'
        ))
        fig_impresiones.add_trace(go.Bar(
            name='Color',
            x=df_dept['organismo'],
            y=df_dept['color'],
            marker_color='lightblue'
        ))
        fig_impresiones.update_layout(
            title='Impresiones B/N vs Color por Departamento',
            barmode='stack',
            xaxis_title='Departamento',
            yaxis_title='Número de Impresiones',
            height=500
        )
        st.plotly_chart(fig_impresiones, use_container_width=True, key="dept_impresiones")
        
        fig_total = px.bar(
            df_dept.sort_values('total_impresiones', ascending=False),
            x='organismo',
            y='total_impresiones',
            title='Total de Impresiones por Departamento',
            labels={'total_impresiones': 'Impresiones Totales', 'organismo': 'Departamento'},
            color='total_impresiones',
            color_continuous_scale='Blues'
        )
        st.plotly_chart(fig_total, use_container_width=True, key="dept_total_imp")
    
    with tab3:
        fig_dispositivos = px.bar(
            df_dept.sort_values('dispositivos', ascending=False),
            x='organismo',
            y='dispositivos',
            title='Número de Dispositivos por Departamento',
            labels={'dispositivos': 'Número de Dispositivos', 'organismo': 'Departamento'},
            color='dispositivos',
            color_continuous_scale='Greens'
        )
        st.plotly_chart(fig_dispositivos, use_container_width=True, key="dept_dispositivos")
        
        df_dept['promedio_por_dispositivo'] = df_dept['total_impresiones'] / df_dept['dispositivos']
        fig_promedio = px.bar(
            df_dept.sort_values('promedio_por_dispositivo', ascending=False),
            x='organismo',
            y='promedio_por_dispositivo',
            title='Promedio de Impresiones por Dispositivo',
            labels={'promedio_por_dispositivo': 'Impresiones/Dispositivo', 'organismo': 'Departamento'},
            color='promedio_por_dispositivo',
            color_continuous_scale='Purples'
        )
        st.plotly_chart(fig_promedio, use_container_width=True, key="dept_promedio")
    
    with tab4:
        df_dept_display = df_dept.copy()
        df_dept_display['coste_sin_iva'] = df_dept_display['coste_sin_iva'].apply(lambda x: f"{x:.2f} €")
        df_dept_display['iva_total'] = df_dept_display['iva_total'].apply(lambda x: f"{x:.2f} €")
        df_dept_display['coste_con_iva'] = df_dept_display['coste_con_iva'].apply(lambda x: f"{x:.2f} €")
        df_dept_display['bn'] = df_dept_display['bn'].apply(lambda x: f"{x:,}")
        df_dept_display['color'] = df_dept_display['color'].apply(lambda x: f"{x:,}")
        df_dept_display['total_impresiones'] = df_dept_display['total_impresiones'].apply(lambda x: f"{x:,}")
        df_dept_display['promedio_por_dispositivo'] = df_dept_display['promedio_por_dispositivo'].apply(lambda x: f"{x:.0f}")
        
        st.dataframe(
            df_dept_display.rename(columns={
                'organismo': 'Departamento',
                'bn': 'B/N',
                'color': 'Color',
                'coste_sin_iva': 'Coste sin IVA',
                                'coste_con_iva': 'Coste con IVA',
                'dispositivos': 'Dispositivos',
                'total_impresiones': 'Total Impresiones',
                'promedio_por_dispositivo': 'Promedio/Dispositivo'
            }),
            use_container_width=True
        )

def mostrar_detalle_equipos(df):
    """Muestra el detalle por equipo"""
    st.markdown("### 🔍 Detalle por Equipo")
    
    col1, col2 = st.columns(2)
    
    with col1:
        departamentos = ['Todos'] + sorted(df['organismo'].unique().tolist())
        dept_seleccionado = st.selectbox("Filtrar por departamento:", departamentos)
    
    with col2:
        if 'documento' in df.columns:
            documentos = ['Todos'] + sorted(df['documento'].unique().tolist())
            doc_seleccionado = st.selectbox("Filtrar por documento:", documentos)
        else:
            doc_seleccionado = 'Todos'
    
    df_filtrado = df.copy()
    
    if dept_seleccionado != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['organismo'] == dept_seleccionado]
    
    if doc_seleccionado != 'Todos' and 'documento' in df.columns:
        df_filtrado = df_filtrado[df_filtrado['documento'] == doc_seleccionado]
    
    # Formatear el dataframe
    df_display = df_filtrado.copy()
    df_display['coste_sin_iva'] = df_display['coste_sin_iva'].apply(lambda x: f"{x:.2f} €")
    df_display['iva_total'] = df_display['iva_total'].apply(lambda x: f"{x:.2f} €")
    df_display['coste_con_iva'] = df_display['coste_con_iva'].apply(lambda x: f"{x:.2f} €")
    
    columnas_mostrar = ['sn', 'organismo', 'ubicacion', 'bn', 'color', 
                        'coste_sin_iva', 'coste_con_iva', 'estado']
    
    if 'documento' in df_display.columns:
        columnas_mostrar.insert(1, 'documento')
    
    columnas_renombrar = {
        'sn': 'S/N',
        'organismo': 'Organismo',
        'documento': 'Documento',
        'ubicacion': 'Ubicación',
        'bn': 'B/N',
        'color': 'Color',
        'coste_sin_iva': 'Coste sin IVA',
                'coste_con_iva': 'Coste con IVA',
        'estado': 'Estado'
    }
    
    def highlight_missing(row):
        return ['background-color: #ffcccc' if row['Estado'] == '⚠️ Faltante en Excel' else '' for _ in row]

    st.dataframe(
        df_display[columnas_mostrar].rename(columns=columnas_renombrar).style.apply(highlight_missing, axis=1),
        use_container_width=True
    )

######## UI STREAMLIT ########
st.title("📑 FactuBAM — Revisión de Facturas")

# Mostrar información de almacenamiento
if st.session_state.historial_documentos:
    st.success(f"✅ {len(st.session_state.historial_documentos)} registro(s) guardado(s) en disco local")
    st.info(f"📁 Ubicación: `{DATA_DIR.absolute()}`")

# Menú de navegación principal
st.markdown("### 🎯 Modo de Visualización")
col_menu1, col_menu2, col_menu3, col_menu4 = st.columns(4)

with col_menu1:
    if st.button("➕ Nuevo Análisis", use_container_width=True, type="primary" if st.session_state.modo_vista == 'nuevo' else "secondary"):
        st.session_state.modo_vista = 'nuevo'
        st.rerun()

with col_menu2:
    if st.button("📄 Ver Documento", use_container_width=True, type="primary" if st.session_state.modo_vista == 'individual' else "secondary"):
        st.session_state.modo_vista = 'individual'
        st.rerun()

with col_menu3:
    if st.button("📊 Vista Acumulada", use_container_width=True, 
                 type="primary" if st.session_state.modo_vista == 'acumulado' else "secondary",
                 disabled=len(st.session_state.historial_documentos) == 0):
        st.session_state.modo_vista = 'acumulado'
        st.rerun()

with col_menu4:
    if st.button("🔄 Comparar Documentos", use_container_width=True,
                 type="primary" if st.session_state.modo_vista == 'comparativa' else "secondary",
                 disabled=len(st.session_state.historial_documentos) < 2):
        st.session_state.modo_vista = 'comparativa'
        st.rerun()

st.markdown("---")

# Sección de gestión de documentos guardados
if st.session_state.historial_documentos and st.session_state.modo_vista != 'nuevo':
    with st.expander("📚 Gestión de Documentos Guardados", expanded=False):
        for registro in st.session_state.historial_documentos:
            col1, col2, col3, col4, col5 = st.columns([3, 2, 1, 1, 1])
            
            with col1:
                nuevo_nombre = st.text_input(
                    "✏️ Renombrar",
                    value=registro['nombre'],
                    key=f"rename_{registro['id']}"
                )
                if nuevo_nombre != registro['nombre']:
                    registro['nombre'] = nuevo_nombre
                    guardar_historial(st.session_state.historial_documentos)
            
            with col2:
                st.text(registro['fecha_hora'])
            
            with col3:
                st.text(f"{registro['dispositivos']} disp.")
            
            with col4:
                st.text(f"{registro['coste_total_con_iva']:.2f} €")
            
            with col5:
                if st.button("🗑️", key=f"del_{registro['id']}"):
                    eliminar_registro(registro['id'])
                    if st.session_state.registro_seleccionado == registro['id']:
                        st.session_state.registro_seleccionado = None
                    st.rerun()
        
        st.markdown("---")
        if st.button("🗑️ Limpiar todo el historial"):
            limpiar_historial()
            st.session_state.registro_seleccionado = None
            st.session_state.modo_vista = 'nuevo'
            st.rerun()

st.markdown("---")

# Renderizar según modo de vista
if st.session_state.modo_vista == 'nuevo':
    st.subheader("➕ Cargar Nuevo Documento")
    
    pdf_file = st.file_uploader("Sube la factura PDF", type=["pdf"])
    excel_file = st.file_uploader("Sube el inventario Excel", type=["xlsx"])

    if pdf_file and excel_file:
        nombre_registro = st.text_input("📝 Nombre para este análisis:", placeholder="Ej: Factura Enero 2024")
        
        if st.button("Procesar y Guardar", type="primary"):
            if nombre_registro:
                with st.spinner("Procesando documentos..."):
                    datos_pdf = extraer_datos_pdf(pdf_file)
                    resultados = cruzar_excel(excel_file, datos_pdf)
                    df = pd.DataFrame(resultados)
                    
                    guardar_registro(nombre_registro, pdf_file, excel_file, df)
                    
                    st.success(f"✅ Análisis '{nombre_registro}' guardado correctamente")
                    st.session_state.modo_vista = 'individual'
                    st.session_state.registro_seleccionado = st.session_state.historial_documentos[-1]['id']
                    st.rerun()
            else:
                st.warning("⚠️ Por favor, ingresa un nombre para el análisis")

elif st.session_state.modo_vista == 'individual':
    st.subheader("📄 Ver Documento Individual")
    
    if st.session_state.historial_documentos:
        nombres_docs = [f"{reg['nombre']} ({reg['fecha_hora']})" for reg in st.session_state.historial_documentos]
        doc_seleccionado_idx = st.selectbox(
            "Selecciona un documento:",
            range(len(st.session_state.historial_documentos)),
            format_func=lambda x: nombres_docs[x]
        )
        
        registro = st.session_state.historial_documentos[doc_seleccionado_idx]
        st.session_state.registro_seleccionado = registro['id']
        
        mostrar_analisis(registro['df'], titulo=f"📊 Análisis: {registro['nombre']}")
    else:
        st.info("No hay documentos guardados. Carga uno nuevo desde el menú 'Nuevo Análisis'.")

elif st.session_state.modo_vista == 'acumulado':
    st.subheader("📊 Vista Acumulada - Todos los Documentos")
    
    st.info(f"📁 Mostrando datos acumulados de {len(st.session_state.historial_documentos)} documento(s)")
    
    # Selector de documentos a incluir
    with st.expander("⚙️ Configurar documentos a incluir", expanded=False):
        st.markdown("**Selecciona qué documentos incluir en el análisis acumulado:**")
        
        if st.checkbox("Seleccionar todos", value=True):
            ids_seleccionados = [reg['id'] for reg in st.session_state.historial_documentos]
        else:
            ids_seleccionados = []
            for registro in st.session_state.historial_documentos:
                if st.checkbox(f"{registro['nombre']} ({registro['fecha_hora']})", 
                               value=False, 
                               key=f"check_acum_{registro['id']}"):
                    ids_seleccionados.append(registro['id'])
    
    if ids_seleccionados:
        df_acumulado = obtener_dataframe_acumulado(ids_seleccionados)
        
        if df_acumulado is not None:
            mostrar_analisis(df_acumulado, 
                           titulo=f"📊 Análisis Acumulado ({len(ids_seleccionados)} documento(s))",
                           mostrar_por_documento=True)
    else:
        st.warning("⚠️ Selecciona al menos un documento para ver el análisis acumulado")

elif st.session_state.modo_vista == 'comparativa':
    st.subheader("🔄 Comparar Documentos")
    
    st.markdown("**Selecciona los documentos que deseas comparar:**")
    
    col_select1, col_select2 = st.columns(2)
    
    with col_select1:
        st.markdown("#### 📄 Documento 1")
        nombres_docs = [f"{reg['nombre']}" for reg in st.session_state.historial_documentos]
        doc1_idx = st.selectbox(
            "Selecciona el primer documento:",
            range(len(st.session_state.historial_documentos)),
            format_func=lambda x: nombres_docs[x],
            key="comp_doc1"
        )
    
    with col_select2:
        st.markdown("#### 📄 Documento 2")
        doc2_idx = st.selectbox(
            "Selecciona el segundo documento:",
            range(len(st.session_state.historial_documentos)),
            format_func=lambda x: nombres_docs[x],
            key="comp_doc2"
        )
    
    if doc1_idx != doc2_idx:
        reg1 = st.session_state.historial_documentos[doc1_idx]
        reg2 = st.session_state.historial_documentos[doc2_idx]
        
        st.markdown("---")
        st.markdown("### 📊 Comparativa de Métricas")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "Dispositivos",
                reg2['dispositivos'],
                delta=reg2['dispositivos'] - reg1['dispositivos'],
                delta_color="normal"
            )
        
        with col2:
            total_imp1 = reg1['df']['bn'].sum() + reg1['df']['color'].sum()
            total_imp2 = reg2['df']['bn'].sum() + reg2['df']['color'].sum()
            st.metric(
                "Total Impresiones",
                f"{total_imp2:,}",
                delta=total_imp2 - total_imp1,
                delta_color="normal"
            )
        
        with col3:
            st.metric(
                "Coste sin IVA",
                f"{reg2['coste_total_sin_iva']:.2f} €",
                delta=f"{reg2['coste_total_sin_iva'] - reg1['coste_total_sin_iva']:.2f} €",
                delta_color="inverse"
            )
        
        with col4:
            st.metric(
                "Coste con IVA",
                f"{reg2['coste_total_con_iva']:.2f} €",
                delta=f"{reg2['coste_total_con_iva'] - reg1['coste_total_con_iva']:.2f} €",
                delta_color="inverse"
            )
        
        st.markdown("---")
        
        # Gráfico comparativo de costes
        df_comparacion = pd.DataFrame({
            'Documento': [reg1['nombre'], reg2['nombre']],
            'Sin IVA': [reg1['coste_total_sin_iva'], reg2['coste_total_sin_iva']],
            'Con IVA': [reg1['coste_total_con_iva'], reg2['coste_total_con_iva']]
        })
        
        fig_comp_costes = go.Figure()
        fig_comp_costes.add_trace(go.Bar(
            name='Sin IVA',
            x=df_comparacion['Documento'],
            y=df_comparacion['Sin IVA'],
            marker_color='lightcoral'
        ))
        fig_comp_costes.add_trace(go.Bar(
            name='Con IVA',
            x=df_comparacion['Documento'],
            y=df_comparacion['Con IVA'],
            marker_color='darkred'
        ))
        fig_comp_costes.update_layout(
            title='Comparación de Costes',
            barmode='group',
            height=400
        )
        st.plotly_chart(fig_comp_costes, use_container_width=True, key="comp_costes_docs")
        
        # Comparación de impresiones
        col_imp1, col_imp2 = st.columns(2)
        
        with col_imp1:
            st.markdown(f"#### {reg1['nombre']}")
            fig1 = px.pie(
                values=[reg1['df']['bn'].sum(), reg1['df']['color'].sum()],
                names=['B/N', 'Color'],
                title='Distribución de Impresiones',
                hole=0.4
            )
            st.plotly_chart(fig1, use_container_width=True, key="comp_pie1")
        
        with col_imp2:
            st.markdown(f"#### {reg2['nombre']}")
            fig2 = px.pie(
                values=[reg2['df']['bn'].sum(), reg2['df']['color'].sum()],
                names=['B/N', 'Color'],
                title='Distribución de Impresiones',
                hole=0.4
            )
            st.plotly_chart(fig2, use_container_width=True, key="comp_pie2")
        
        # Análisis por departamento comparativo
        st.markdown("---")
        st.markdown("### 🏢 Comparación por Departamentos")
        
        df1_dept = reg1['df'].groupby('organismo').agg({
            'coste_con_iva': 'sum'
        }).reset_index()
        df1_dept['documento'] = reg1['nombre']
        
        df2_dept = reg2['df'].groupby('organismo').agg({
            'coste_con_iva': 'sum'
        }).reset_index()
        df2_dept['documento'] = reg2['nombre']
        
        df_dept_comp = pd.concat([df1_dept, df2_dept])
        
        fig_dept = px.bar(
            df_dept_comp,
            x='organismo',
            y='coste_con_iva',
            color='documento',
            barmode='group',
            title='Coste por Departamento (con IVA)',
            labels={'coste_con_iva': 'Coste (€)', 'organismo': 'Departamento'}
        )
        st.plotly_chart(fig_dept, use_container_width=True, key="comp_dept_costes")
        
    else:
        st.warning("⚠️ Por favor, selecciona dos documentos diferentes para comparar")

# ======================================================
# UTILIDADES MD5 – DETECCIÓN DE ARCHIVOS DUPLICADOS
# ======================================================
from collections import defaultdict

def calcular_md5_archivo(ruta_archivo, bloque_size=8192):
    """Calcula el MD5 de un archivo leyendo por bloques"""
    md5 = hashlib.md5()
    with open(ruta_archivo, "rb") as f:
        for bloque in iter(lambda: f.read(bloque_size), b""):
            md5.update(bloque)
    return md5.hexdigest()

def detectar_duplicados_md5():
    """Detecta archivos duplicados por MD5 en factubam_data/documentos"""
    hashes = defaultdict(list)

    for archivo in DOCUMENTOS_DIR.iterdir():
        if archivo.is_file():
            try:
                md5 = calcular_md5_archivo(archivo)
                hashes[md5].append(archivo.name)
            except Exception as e:
                print(f"Error leyendo {archivo.name}: {e}")

    duplicados = {md5: files for md5, files in hashes.items() if len(files) > 1}
    return duplicados
